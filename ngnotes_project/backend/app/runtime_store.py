"""
NGNotes: In-memory runtime data store and Excel export.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from threading import Lock
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class RuntimeStore:
    """Thread-safe in-memory store for runtime rows."""

    def __init__(self) -> None:
        self._rows: List[Dict[str, Any]] = []
        self._lock = Lock()

    def add_row(self, row: Dict[str, Any]) -> None:
        with self._lock:
            self._rows.append(row)

    def add_rows(self, rows: List[Dict[str, Any]]) -> None:
        with self._lock:
            self._rows.extend(rows)

    def clear(self) -> int:
        with self._lock:
            deleted = len(self._rows)
            self._rows = []
            return deleted

    def count(self) -> int:
        with self._lock:
            return len(self._rows)

    def list_rows(self) -> List[Dict[str, Any]]:
        with self._lock:
            return list(self._rows)

    def export_excel_bytes(self) -> bytes:
        rows = self.list_rows()

        wb = Workbook()
        ws = wb.active
        ws.title = "runtime_data"

        headers = [
            "temperature",
            "top_p",
            "min_p",
            "top_k",
            "max_tokens",
            "repetition_penalty",
            "rouge_l_f1",
            "semantic_similarity",
            "composite_score",
            "rubric_total_score",
            "hallucination_score",
            "context_adherence",
            "domain_fluency",
            "final_score",
        ]

        ws.append(headers)

        for row in rows:
            ws.append([row.get(h) for h in headers])

        # Light quality-of-life formatting.
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{max(2, len(rows) + 1)}"

        output = BytesIO()
        wb.save(output)
        return output.getvalue()


def now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
